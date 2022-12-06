import os
import pytesseract
from PIL import Image
from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtWidgets import QFileDialog
import numpy as np
import time
import shutil
from openpyxl import load_workbook
import cv2
import winreg

import match

# 局部变量，图片路径，配置，返回结果
image_path = 'image_path'
def_excel_path = 'def_excel_path'
program_path = 'program_path'

image_name = []  # 图像名字
osc_value_r = []  # 纹波名和值
osc_value_n = []  # 噪声名和值

config_get = []  # 获取配置信息
negative_film = True
image_matching = True
fast_recognition = False
super_fast_recognition = True

RS_new_jpg = 'user/template/RS_new_jpg.png'
RS_old_jpg = 'user/template/RS_old_jpg.png'  # 未有
RS_new_png = 'user/template/RS_new_png.png'
RS_old_png = 'user/template/RS_old_png.png'
TK_jpg = 'user/template/TK_jpg.png'  # 未有
TK_png = 'user/template/TK_png.png'

rs_lib = 'user/template/rs_words_lib.png'  # RS匹配图像库

tesseract_path = 'Tesseract-OCR/tesseract.exe'  # Tesseract路径
excel_path = 'user/osc_value.xlsx'  # 输出模板路径


def Tesseract_OCR(image, orc_lang):
    pytesseract.pytesseract.tesseract_cmd = program_path + tesseract_path
    osc_value = pytesseract.image_to_string(image, lang=orc_lang)  # 识别图片文字
    return osc_value


def path_to_python_style(path):
    if path.find('\\') != -1:  # 找到使用\的
        path = path.replace('\\', '/')  # 替换为/
    if path[-1] != '/':
        path += '/'
    return path


class Worker(QThread):
    signal1 = pyqtSignal(int, str)  # 定义信号类型

    def __init__(self):  # 不知道有啥用？
        super(Worker, self).__init__()

    def osc_value_program(self):  # 核心识别代码
        global image_path, image_name

        time_start = time.time()  # 获取当前开始时间

        # 获取处理截图文件夹下所有文件名列表，并区分出图片扩展名
        image_path_list = os.listdir(image_path)
        # print(list)

        for img in image_path_list:
            file_ext = os.path.splitext(image_path + img)[1]
            if file_ext == '.jpg' or file_ext == '.JPG' or file_ext == '.png' or file_ext == '.PNG':
                image_name.append(img)
        # print(image_name)

        # 主循环
        image_name_index = 0
        for img_name in image_name:
            image_load = Image.open(image_path + img_name)  # 打开图片
            orc_lang = 'rs'
            offset = 103  # RS_new_jpg 103, RS_old_png 103, RS_new_png 103, TK_jpg 67
            if image_load.format == 'PNG':  # 获取图片信息，以区分三台示波器，区分jpg/png
                if image_load.info.get('dpi') == (96, 96):
                    crop_osc = config_get[1]  # crop_TK_png
                    template_image = TK_png
                    orc_lang = 'tk'
                    offset = 67
                else:
                    if image_load.info.get('InstrumentSerialNumber') == '320026':
                        crop_osc = config_get[3]  # crop_RS_new_png
                        template_image = RS_new_png
                    else:
                        crop_osc = config_get[5]  # crop_RS_old_png
                        template_image = RS_old_png
            else:
                if image_load.info.get('dpi') == (96, 96):
                    crop_osc = config_get[0]  # crop_TK_jpg
                    template_image = TK_jpg
                    orc_lang = 'tk'
                    offset = 68
                else:
                    if image_load.size == (1280, 800):
                        crop_osc = config_get[2]  # crop_RS_new_jpg
                        template_image = RS_new_jpg
                    else:
                        crop_osc = config_get[4]  # crop_RS_old_jpg
                        template_image = RS_old_jpg
            # print(template_image)

            # 解决中文路径问题# 图像灰度处理cv2打开图片# 读取目标图片
            image_cv = cv2.imdecode(np.fromfile(image_path + img_name, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)

            # cv2.imshow('Image', image_cv)  # 显示图像
            # cv2.waitKey(0)  # 0为不限时间,接收到键盘指令才会继续。也可以输入正整数，单位为毫秒。

            if image_matching:  # 判断是否需要采用图像匹配方法
                template = cv2.imdecode(np.fromfile(program_path + template_image, dtype=np.uint8),
                                        cv2.IMREAD_GRAYSCALE)  # 读取模板图片

                if not negative_film:  # 判断是否需要反相处理
                    template = cv2.bitwise_not(template)

                # cv2.imshow('Image', template)  # 显示图像
                # cv2.waitKey(0)  # 0为不限时间,接收到键盘指令才会继续。也可以输入正整数，单位为毫秒。

                theight, twidth = template.shape[:2]  # 获得模板图片的高宽尺寸
                result = cv2.matchTemplate(image_cv, template, cv2.TM_SQDIFF_NORMED)  # 执行模板匹配，采用的匹配方式
                cv2.normalize(result, result, 0, 1, cv2.NORM_MINMAX, -1)  # 归一化处理
                # 寻找矩阵（一维数组当做向量，用Mat定义）中的最大值和最小值的匹配结果及其位置
                _, _, min_loc, _ = cv2.minMaxLoc(result)
                x0 = min_loc[0] + twidth  # 接着模板后面就是数据
                y0 = min_loc[1]
                x1 = x0 + offset
                y1 = y0 + theight - 1
                image_cv = image_cv[y0:y1, x0:x1]  # 裁剪坐标为[y0:y1,x0:x1]
            else:
                image_cv = image_cv[crop_osc[1]:crop_osc[3], crop_osc[0]:crop_osc[2]]  # 裁剪坐标为[y0:y1,x0:x1]

            image_cv = cv2.GaussianBlur(image_cv, (1, 1), 0)  # 高斯模糊处理

            image_cv = cv2.resize(image_cv, None, fx=5, fy=5, interpolation=cv2.INTER_CUBIC)  # 对图像放大5倍

            if negative_film:  # 判断是否需要反相处理
                my_threshold = cv2.THRESH_BINARY_INV
            else:
                my_threshold = cv2.THRESH_BINARY
            # OTSU大津算法自适应阈值二值化
            _, image_cv = cv2.threshold(image_cv, 0, 255, cv2.THRESH_OTSU | my_threshold)

            # cv2.imshow('Image', image_cv)  # 显示图像
            # cv2.waitKey(0)  # 0为不限时间,接收到键盘指令才会继续。也可以输入正整数，单位为毫秒。

            if orc_lang == 'tk':
                if not fast_recognition:  # 采用标准识别英文文件
                    orc_lang = 'eng'
                osc_value = Tesseract_OCR(image_cv, orc_lang)  # 识别图片文字
            else:
                if not super_fast_recognition:
                    if not fast_recognition:  # 采用标准识别英文文件
                        orc_lang = 'eng'
                    osc_value = Tesseract_OCR(image_cv, orc_lang)  # 识别图片文字
                else:
                    match.words_lib = program_path + rs_lib
                    osc_value = match.osc_matching(image_cv)  # 采用匹配识别方式

            # print(osc_value)

            # cv2.imshow('Image', image_cv)  # 显示图像
            # cv2.waitKey(0)  # 0为不限时间,接收到键盘指令才会继续。也可以输入正整数，单位为毫秒。
            # cv2.destroyAllWindows()  # 关闭我们所有cv的窗口

            # 处理识别文本
            try:
                osc_value_end = osc_value.index('m')  # 查找其中的‘m’
                osc_value_mag = 1  # 倍率，以毫伏为单位
            except ValueError:
                try:
                    osc_value_end = osc_value.index('u')  # 查找其中的‘u',识别问题
                    osc_value_mag = 0.001
                except ValueError:
                    try:
                        osc_value_end = osc_value.index('p')  # 查找其中的‘p’,识别问题
                        osc_value_mag = 0.001
                    except ValueError:
                        try:
                            osc_value_end = osc_value.index('¥')  # 查找其中的‘¥’，识别问题
                            osc_value_mag = 1000
                        except ValueError:
                            try:
                                osc_value_end = osc_value.index('V')  # 查找其中的‘V’
                                osc_value_mag = 1000
                            except ValueError:
                                try:
                                    osc_value_end = osc_value.index('v')  # 查找其中的‘v’
                                    osc_value_mag = 1000
                                except ValueError:
                                    self.signal1.emit(-1, "'%s' OCR 'mp¥Vv' ERROR!\n" % img_name)  # 识别错误吗，发送识别信息
                                    # break
                                    osc_value_end = 1
                                    osc_value_mag = 10000

            osc_value = osc_value[:osc_value_end]  # 分割出数值
            osc_value = osc_value.replace(' ', '')  # 去空格
            try:
                osc_value_float = float(osc_value) * osc_value_mag  # 转换成浮点型，单位mV
            except ValueError:
                self.signal1.emit(-1, "'%s' OCR 'float' ERROR!\n" % img_name)  # 转换错误，发送错误信息
                osc_value_float = 999999  # 标志错误值

            # 处理文件名，与数据一一对应，区分纹波噪声
            if img_name.upper().find("-R.") != -1:
                osc_value_r.append([img_name[:-6].upper(), osc_value_float])
            elif img_name.upper().find("-N.") != -1:
                osc_value_n.append([img_name[:-6].upper(), osc_value_float])

            p = int((image_name_index + 1) / len(image_name) * 100)
            log_print = img_name[:-6].upper() + (' %.3f' % osc_value_float) + 'mV\n'
            self.signal1.emit(p, log_print)  # 发送进度条的值 信号

            image_name_index += 1

        time_stop = time.time()  # 结束时间
        self.signal1.emit(101, ("%d pictures. Time consuming is %.*fs.\n" % (
            len(image_name), 3, time_stop - time_start)))  # 发送处理耗时

    def osc_value_save_to_excel(self):
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        save_path = winreg.QueryValueEx(key, "Desktop")[0]  # 获取桌面目录

        file_name_path = QFileDialog.getSaveFileName(None, "文件保存", save_path, 'Excel文件 (*.xlsx)')  # 获取保存路径
        new_excel_path = file_name_path[0]

        if new_excel_path is '':  # 点击取消
            self.signal1.emit(-1, "Data saving path not selected, data not saved, please start again!\n")
            return

        excel_name = new_excel_path[new_excel_path[:-1].rfind("/") + 1:]  # 获取保存文件名

        try:
            shutil.copyfile(program_path + excel_path, new_excel_path)  # 将excel模板复制到图片文件夹下
        except IOError:  # 复制识别，目标文件被打开或者源文件不存在
            self.signal1.emit(-1, "The file '%s' is already open, or the template file under the '%s' path does not "
                                  "exist, please check and try again!\n" % (excel_name, program_path + excel_path))
            return

        workbook = load_workbook(def_excel_path)  # 打开标准文件
        workbook = workbook.active  # 激活表格

        row_e = []
        row_e_name = []
        for row in workbook.iter_rows(min_row=3, min_col=1, max_col=4, values_only=True):  # 获取标准数据
            if row == (None, None, None, None):
                break
            row_l = list(row)  # 组元转换成列表
            row_e.append(row_l)  # 所有数据
            row_e_name.append(row_l[1].upper())  # 测试点大写处理

        workbook_s = load_workbook(new_excel_path)  # 打开excel并写入数据
        sheet_ranges = workbook_s['R_N']  # 激活名为‘R_N’的表格

        osc_value_r_name = np.array(osc_value_r)  # 二维列表转化为矩阵
        osc_value_r_name = osc_value_r_name[:, 0]  # 取第0列，纹波名
        osc_value_r_name = osc_value_r_name.tolist()  # 一维矩阵转化为列表

        osc_value_n_name = np.array(osc_value_n)  # 二维列表转化为矩阵
        osc_value_n_name = osc_value_n_name[:, 0]  # 取第0列，噪声名
        osc_value_n_name = osc_value_n_name.tolist()  # 一维矩阵转化为列表

        missing_test_r = []  # 获取标准没有匹配上的名字
        missing_test_n = []  # 获取标准没有匹配上的名字
        no_match_image = []  # 获取图像没有匹配上的名字

        for y in range(len(row_e)):
            for x in range(4):
                sheet_ranges.cell(row=y + 3, column=x + 1, value=row_e[y][x])  # 写入标准

            try:
                osc_value_r_t_index = osc_value_r_name.index(row_e_name[y])  # 获取标准匹配上纹波的序号
                sheet_ranges.cell(row=y + 3, column=10, value=osc_value_r[osc_value_r_t_index][0])  # 写入纹波测试点名
                sheet_ranges.cell(row=y + 3, column=5, value=osc_value_r[osc_value_r_t_index][1])  # 写入纹波数据
            except ValueError:
                missing_test_r.append(row_e_name[y])  # 获取标准没有匹配上的名字

            try:
                osc_value_n_t_index = osc_value_n_name.index(row_e_name[y])  # 获取标准匹配上噪声的序号
                sheet_ranges.cell(row=y + 3, column=11, value=osc_value_n[osc_value_n_t_index][0])  # 写入噪声测试点名
                sheet_ranges.cell(row=y + 3, column=6, value=osc_value_n[osc_value_n_t_index][1])  # 写入噪声数据
            except ValueError:
                missing_test_n.append(row_e_name[y])  # 获取标准没有匹配上的名字

        workbook_s.save(new_excel_path)  # 保存excel
        os.startfile(new_excel_path)  # 打开excel

        for image in image_name:
            if image[:-6].upper() not in row_e_name:
                no_match_image.append(image)  # 获取图像没有匹配上的名字

        error_message = ''
        if no_match_image:
            error_message += '这些截图没有对应的测试点: '  # 获取图像没有匹配上的名字
            for name in no_match_image:
                error_message += name + ', '
            error_message = error_message[:-2] + '\n\n'

        if missing_test_r:
            error_message += '这些波纹测试点没有对应的截图: '  # 获取标准没有匹配上的名字
            for name in missing_test_r:
                error_message += name + ', '
            error_message = error_message[:-2] + '\n\n'

        if missing_test_n:
            error_message += '这些噪声测试点没有对应的截图: '  # 获取标准没有匹配上的名字
            for name in missing_test_n:
                error_message += name + ', '
            error_message = error_message[:-2] + '\n\n'

        if no_match_image or missing_test_r or missing_test_n:
            self.signal1.emit(-1, error_message)

    def run(self):
        global image_path, program_path

        image_path = path_to_python_style(image_path)
        program_path = path_to_python_style(program_path)

        self.osc_value_program()

        if osc_value_r and osc_value_n:  # 识别数据不为空
            self.osc_value_save_to_excel()  # 写入Excel

        osc_value_r.clear()  # 清空数据
        osc_value_n.clear()
        config_get.clear()
        image_name.clear()
